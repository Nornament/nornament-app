"""The shape of the IVY Karigar stock export, in one place.

Products are blocks, not rows: a block opens wherever column C carries a style
number and runs to the next one. Five material bands then run down that block
*independently* — the stone on the third row of a block has nothing to do with
the diamond on the first. Reading them in lockstep would silently pair up
unrelated materials, which is why each band gets its own pass.
"""
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as col

SHEET = "Sheet1"
HEADER_ROW = 3
FIRST_DATA_ROW = 4

#: Column letters we check in row 3 before trusting anything else in the file.
EXPECTED_HEADERS = {
    "C": "Style No",
    "D": "JewelCode",
    "E": "Category",
    "M": "Collection",
    "U": "Item Code",
    "AK": "Item Code",
    "AV": "Stone",
    "BE": "Cost Price",
    "BS": "Item Code",
}

#: band name -> the columns that make up one line of that band.
#: ``key`` is the column that decides whether the band has a line on this row.
BANDS = {
    "diamond": {
        "key": "U", "name": "V", "shape": "X", "quality": "Y",
        "size_band": "AA", "pcs": "AD", "qty": "AE",
        "cost_rate": "AG", "sale_rate": "AH",
        "cost_amount": "AI", "sale_amount": "AJ",
    },
    "metal": {
        "key": "AK", "name": "AL", "qty": "AN",
        "cost_rate": "AR", "sale_rate": "AS",
        "cost_amount": "AT", "sale_amount": "AU",
    },
    "stone": {
        "key": "AV", "name": "AW", "pcs": "AX", "qty": "AY",
        "cost_rate": "BA", "sale_rate": "BB",
        "cost_amount": "BC", "sale_amount": "BD",
    },
    # BI–BR carries a name but no item code, so the name is the key
    "other": {
        "key": "BJ", "name": "BJ", "qty": "BN",
        "cost_rate": "BO", "sale_rate": "BQ",
        "cost_amount": "BP", "sale_amount": "BR",
    },
    "charge": {
        "key": "BS", "name": "BT", "pcs": "BU", "qty": "BV",
        "cost_rate": "BW", "sale_rate": "BY",
        "cost_amount": "BX", "sale_amount": "BZ",
    },
}


@dataclass
class ParsedLine:
    band: str
    code: str
    name: str = ""
    pcs: int = None
    qty: Decimal = None
    cost_rate: Decimal = None
    sale_rate: Decimal = None
    cost_amount: Decimal = None
    sale_amount: Decimal = None
    size_band: str = ""
    shape: str = ""
    quality: str = ""


@dataclass
class ParsedPiece:
    row_no: int
    sr_no: str = ""
    style_code: str = ""
    jewel_code: str = ""
    category: str = ""
    sub_category: str = ""
    collection: str = ""
    vendor: str = ""
    make_type: str = ""
    stock_type: str = ""
    inw_date: object = None
    fg_date: object = None
    metal_purity: str = ""
    diamond_quality: str = ""
    remarks: str = ""
    src_cost_price: Decimal = None
    src_sale_price: Decimal = None
    src_net_wt_gm: Decimal = None
    making_cost: Decimal = None
    making_sale: Decimal = None
    lines: list = field(default_factory=list)
    image: bytes = None


def _text(sheet, row, letter):
    value = sheet.cell(row=row, column=col(letter)).value
    if value is None:
        return ""
    return str(value).strip()


def _num(sheet, row, letter):
    """A cell as Decimal, or None. Blank and unparseable both mean None."""
    value = sheet.cell(row=row, column=col(letter)).value
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _int(sheet, row, letter):
    value = _num(sheet, row, letter)
    return None if value is None else int(value)


def _date(value):
    """The export writes dates two ways: real datetimes, and 'Nov  7 2025 12:00AM'."""
    if isinstance(value, datetime):
        return value.date()
    if not value:
        return None
    text = " ".join(str(value).split())
    for pattern in ("%b %d %Y %I:%M%p", "%b %d %Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def header_problems(fileobj):
    """Every header in row 3 that is not what this importer expects.

    Checked before anything else, so the wrong workbook is refused with the
    offending column rather than imported as nonsense. Anything that is not a
    readable workbook, or is one without the export's sheet, is a problem too
    — this is the gate for files a person picked by mistake, so it has to
    answer rather than raise.
    """
    try:
        book = load_workbook(fileobj, data_only=True, read_only=True)
    except Exception as error:
        return [f"That file could not be read as a spreadsheet ({error})."]
    if SHEET not in book.sheetnames:
        return [f"No {SHEET!r} sheet — found {', '.join(book.sheetnames) or 'nothing'}."]
    sheet = book[SHEET]
    problems = []
    for letter, expected in EXPECTED_HEADERS.items():
        found = _text(sheet, HEADER_ROW, letter)
        if found != expected:
            problems.append(f"Column {letter} should be {expected!r}, found {found!r}")
    return problems


def _line_at(sheet, row, band, spec):
    """One line of one band on one row, or None if this band is blank here."""
    code = _text(sheet, row, spec["key"])
    if not code:
        return None
    return ParsedLine(
        band=band,
        code=code,
        name=_text(sheet, row, spec.get("name", spec["key"])),
        pcs=_int(sheet, row, spec["pcs"]) if "pcs" in spec else None,
        qty=_num(sheet, row, spec["qty"]) if "qty" in spec else None,
        cost_rate=_num(sheet, row, spec["cost_rate"]),
        sale_rate=_num(sheet, row, spec["sale_rate"]),
        cost_amount=_num(sheet, row, spec["cost_amount"]),
        sale_amount=_num(sheet, row, spec["sale_amount"]),
        size_band=_text(sheet, row, spec["size_band"]) if "size_band" in spec else "",
        shape=_text(sheet, row, spec["shape"]) if "shape" in spec else "",
        quality=_text(sheet, row, spec["quality"]) if "quality" in spec else "",
    )


def _images_by_row(sheet):
    """Anchor row -> image bytes. Images sit in column A on parent rows."""
    found = {}
    for image in getattr(sheet, "_images", []):
        row = image.anchor._from.row + 1
        data = image._data() if callable(getattr(image, "_data", None)) else None
        if data:
            found[row] = data
    return found


def parse(fileobj):
    """Every product in the workbook, with its material lines and its photo."""
    book = load_workbook(fileobj, data_only=True)
    sheet = book[SHEET]
    images = _images_by_row(sheet)

    starts = [
        row for row in range(FIRST_DATA_ROW, sheet.max_row + 1)
        if _text(sheet, row, "C")
    ]
    pieces = []
    for index, start in enumerate(starts):
        end = starts[index + 1] if index + 1 < len(starts) else sheet.max_row + 1
        piece = ParsedPiece(
            row_no=start,
            sr_no=_text(sheet, start, "B"),
            style_code=_text(sheet, start, "C"),
            jewel_code=_text(sheet, start, "D"),
            category=_text(sheet, start, "E"),
            sub_category=_text(sheet, start, "F"),
            collection=_text(sheet, start, "M"),
            vendor=_text(sheet, start, "K"),
            make_type=_text(sheet, start, "O"),
            stock_type=_text(sheet, start, "P").upper().replace(" ", "_"),
            inw_date=_date(sheet.cell(row=start, column=col("H")).value),
            fg_date=_date(sheet.cell(row=start, column=col("Q")).value),
            metal_purity=_text(sheet, start, "BI"),
            diamond_quality=_text(sheet, start, "Y"),
            remarks=_text(sheet, start, "R"),
            src_cost_price=_num(sheet, start, "BE"),
            src_sale_price=_num(sheet, start, "BF"),
            src_net_wt_gm=_num(sheet, start, "AN"),
            # BG/BH are the piece's making charge. They sit with the totals
            # rather than in a band, so they are read here and become a labour
            # line at commit — without them every piece imports under-costed.
            making_cost=_num(sheet, start, "BG"),
            making_sale=_num(sheet, start, "BH"),
            image=images.get(start),
        )
        # each band down the whole block, on its own — see the module docstring
        for band, spec in BANDS.items():
            for row in range(start, end):
                line = _line_at(sheet, row, band, spec)
                if line is not None:
                    piece.lines.append(line)
        pieces.append(piece)
    return pieces
