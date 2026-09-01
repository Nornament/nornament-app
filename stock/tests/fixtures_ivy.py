"""A miniature IVY workbook, built in memory.

Three products, because the fixture has to stay readable: one plain row, one
whose bands are deliberately ragged, and one that collides with a piece the
test has already created.
"""
import io
from datetime import datetime

from openpyxl import Workbook

#: header row 3, verbatim from the real export
HEADERS = [
    "Image", "Sr No", "Style No", "JewelCode", "Category", "Sub Category",
    "Location Name", "Inw Date", "Qty", "Item Pieces", "Manuf. Name",
    "Manufacturer No", "Collection", "Item Size", "Make Type", "Stock Type",
    "Misc Remarks", "Remarks", "Update By", "Update Date",
    "Item Code", "Item Name Diamond", "Shape", "ShapeName", "Quality", "GSize",
    "Size", "Size MM", "Setting", "Pcs", "Wt", "Batch No", "Cost Rate",
    "Sale Rate", "Cost", "Sale",
    "Item Code", "Item Name Gold", "Gross Wt", "Net Wt", "Loss Wt", "Pure Wt",
    "CPFRate", "Cost Rate", "Sale Rate", "Cost", "Sale",
    "Stone", "Stone (Name)", "Pcs", "Wt", "Batch No", "Cost Rate", "Rate",
    "Cost", "Amount",
    "Cost Price", "Sale Price", "Cost Making Amt", "Sale Making Amt",
    "G Qly", "Item Name", "Quality", "ToneCode", "Gross Wt", "Net Wt",
    "Cost Rate", "Cost Amt", "Sale Rate", "Sale Amt",
    "Item Code", "Item Name", "Pcs", "Wt", "Cost Rate", "Cost Amt", "Rate",
    "Amt",
]


def _header(sr, style, jewel, category, collection):
    """Columns B..T of a parent row, as a dict of column index -> value."""
    return {
        2: sr, 3: style, 4: jewel, 5: category, 6: "STUDS",
        8: datetime(2026, 7, 8), 9: 1, 10: 1, 11: "Infinity Venture [IVY]",
        13: collection, 15: "Casting", 16: "Finish Goods",
        17: "Nov  7 2025 12:00AM", 19: "Krisha Gada",
        20: datetime(2026, 7, 27, 10, 19, 44),
    }


#: (header cells, [extra cells per row of the block])
THREE_PRODUCTS = [
    # 1. one row: one diamond line, one metal line, and the totals
    (
        _header(1, "ER00502", "24P00088", "Earring", "Norna"),
        [{
            21: "DRFGH VS-SI", 22: "DiamondRFGH VS-SI", 23: "R", 24: "Round",
            25: "FGH VS-SI", 26: "+2-6.5", 27: "+3", 28: "1.40",
            30: 2, 31: 0.02, 33: 30000, 34: 250000, 35: 600, 36: 5000,
            37: "G18K", 38: "Gold18K", 39: 1.41, 40: 1.406, 42: 1.059,
            43: 2500, 44: 5460, 45: 11291.09, 46: 7677, 47: 15875,
            57: 9683, 58: 29514.56, 59: 1406, 60: 3515, 61: "18K",
            71: "EC", 72: "Extra Charges", 78: 5400,
        }],
    ),
    # 2. ragged bands: 3 diamond lines, 1 metal line, 2 stone lines.
    #    The stone band must keep reading after the metal band has run out.
    (
        _header(2, "ER00409", "24P00111", "Earring", "Lumina"),
        [
            {
                21: "DRIJ SI-I", 22: "DiamondRIJ SI-I", 23: "R", 24: "Round",
                25: "IJ SI-I", 27: "+1", 30: 10, 31: 0.07,
                33: 9500, 34: 35000, 35: 665, 36: 2450,
                37: "G14K", 38: "Gold14K", 39: 9.75, 40: 8.228, 42: 4.822,
                43: 1500, 44: 4319, 45: 8765.45, 46: 35537, 47: 72122,
                48: "SP01C", 49: "Stone Semi Precious01C", 50: 10, 51: 14,
                53: 100, 54: 300, 55: 1400, 56: 4200,
                57: 84438, 58: 227427.18, 59: 8558, 60: 12837, 61: "14K",
            },
            {
                21: "FPL", 22: "Foil Polki", 26: "+12-18", 27: "+12-18",
                30: 4, 31: 1.86, 33: 6000, 34: 21000, 35: 11160, 36: 39060,
                48: "PS01W", 49: "7 MM Moti", 50: 2, 51: 44.5,
                53: 20, 54: 90, 55: 890, 56: 4005,
            },
            {
                21: "DRFGH SI-I", 22: "DiamondRFGH SI-I", 23: "R",
                24: "Round", 25: "FGH SI-I", 27: "00-0",
                30: 81, 31: 0.39, 33: 10300, 34: 35000, 35: 4017, 36: 13650,
            },
        ],
    ),
    # 3. collides with a piece the test creates beforehand
    (
        _header(3, "RG00113", "24P00095", "Ring", "Brilliance"),
        [{
            21: "DRGH VVS-VS", 22: "DiamondRGH VVS-VS", 23: "R", 24: "Round",
            25: "GH VVS-VS", 30: 24, 31: 0.46,
            33: 20000, 34: 86700, 35: 9200, 36: 39882,
            37: "G14K", 38: "Gold14K", 39: 1.22, 40: 1.128, 42: 0.661,
            43: 1500, 44: 4689, 45: 8765.45, 46: 5289, 47: 9887,
            57: 15617, 58: 51407.77, 59: 1128, 60: 1692, 61: "14K",
        }],
    ),
]


def build_workbook(products=None):
    """The three-product workbook as a BytesIO, laid out like the real export."""
    products = THREE_PRODUCTS if products is None else products
    book = Workbook()
    sheet = book.active
    sheet.title = "Sheet1"
    sheet["A1"] = "IVY Karigar Private Limited"
    for index, name in enumerate(HEADERS, start=1):
        sheet.cell(row=3, column=index, value=name)

    row = 4
    for header, block in products:
        for offset, extras in enumerate(block):
            cells = dict(extras)
            if offset == 0:
                cells.update(header)
            for column, value in cells.items():
                sheet.cell(row=row, column=column, value=value)
            row += 1
    sheet.cell(row=row + 1, column=1, value="[admin] : 27:07:2026 11:18")

    stream = io.BytesIO()
    book.save(stream)
    stream.seek(0)
    return stream
